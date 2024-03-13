import contextlib
import logging
import os, re

#from typing import Union
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class DataProcessor:

  def __init__(self):
    self.input_file = "west.xlsx"
    self.output_file = None
    self.list_color_cells = []
    self.list_duplicates = []
    self.list_valid_floorPlans = []
    self.dict_floorPlans = {}

# -------------------------------------------------
#start set up functions-------------------------------------------------
# -------------------------------------------------

  def save_output_file(self):
    print(f"Beginning the determine_output_files(): {self.output_file}")
    output_folder = 'Outputs/'
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_extension = os.path.splitext(self.input_file)[1].lower()
    file_type = {
        'csv': 'csv',
        'xlsx': 'ss',
        'txt': 'txt',
        'log': 'logs'
    }.get(file_extension, 'ss')

    output_file_name = f"output_{current_datetime}.xlsx"
    self.output_file = os.path.join(output_folder, file_type, output_file_name)

    # Check if the output_file already exists
    if os.path.exists(self.output_file):
      # Add a copy number if the file already exists
      copy_number = 1
      while os.path.exists(self.output_file):
        output_file_name = f"output_{current_datetime}_copy{copy_number}.xlsx"
        self.output_file = os.path.join(output_folder, file_type,output_file_name)
        copy_number += 1

    self.wb.save(self.output_file)
    print(f"Ending the save_output_file(): {self.output_file}")
    print(self.output_file)
    return None

  def load_Data(self):
    # This function loads data from the excel sheet and adds a column for issues
    # The "Issues" column will be the first column (column 1)

    self.wb = load_workbook(self.input_file, read_only=False, data_only=False)
    self.ws = self.wb.active  # This part picks the first sheet on the excel

    # Call update_headers() to remove "Department_ID" and update other headers
    self.update_headers()

    # Add the "Issues" column at the same column position as "Department_ID"
    self.issueColumn = 6  # Column F
    self.ws.cell(row=2, column=self.issueColumn).value = "Issues"

    self.initialize_column_index_map()  # Run after adding column for issues
    for c in self.ws["CD"]:
      c.value = ""
    for x in range(3, self.ws.max_row + 1):
      self.ws.cell(row=x, column=self.issueColumn).value = ""
      self.populate_floorPlans([self.ws.cell(row=x, column=self.column_index_map["Flr Pln N"]),
                               self.ws.cell(row=x, column=self.column_index_map["Flr Pln L"]),
                               self.ws.cell(row=x, column=self.column_index_map["Flr Pln D"])])
    return None

  def update_headers(self):
    header_row = self.ws[1]
    column_names = [cell.value for cell in header_row]
    self.column_index_map = {
        name: index + 1
        for index, name in enumerate(column_names)
    }

  def initialize_column_index_map(self):
    # This function initializes a dictionary to map column names to column indices
    header_row = self.ws[2]
    column_names = [cell.value for cell in header_row]
    self.column_index_map = {
        name: index + 1
        for index, name in enumerate(column_names)
    }

    print("Header Row: ", header_row)
    print("Column Index Map:", self.column_index_map)

  def setup_logging(self):
    #this function sets up logging, and sets up the directory to log data to.
    output_folder = 'Outputs/'
    output_directory = os.path.join(output_folder, 'logs')
    os.makedirs(output_directory, exist_ok=True)
    log_file = os.path.join('Outputs', 'logs', 'data_processor.log')
    logging.basicConfig(
        filename=log_file,  # Set the log file name
        level=logging.INFO,  # Set the logging level to INFO
        format='%(asctime)s - %(levelname)s - %(message)s')  # format 3

  def highlight_rows(self):
    #this function loops through the data and highlight the rows with issues
    #then it loops through the cell issues list and highlight those cells
    yellow = "00FFFF66"
    orange = "00FF9933"
    red = "00FF0000"
    pink = "00FF66FF"
    for i in range(3, self.ws.max_row + 1):
      col = self.ws.cell(row=i, column=self.issueColumn).value
      if len(col.split("/")) == 2:
        for n in range(1, 82):
          x = self.ws.cell(row=i, column=n)
          x.fill = PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
      elif len(col.split("/")) == 3:
        for n in range(1, 82):
          x = self.ws.cell(row=i, column=n)
          x.fill = PatternFill(start_color=orange,end_color=orange,fill_type="solid")
      elif len(col.split("/")) >= 4:
        for n in range(1, 82):
          x = self.ws.cell(row=i, column=n)
          x.fill = PatternFill(start_color=red,end_color=red,fill_type="solid")
      else:
        pass

    for i in self.list_color_cells:
      i.fill = PatternFill(start_color=pink, end_color=pink, fill_type="solid")
    return None

  def populate_floorPlans(self, data):
    #this function populates the floor plan dictionary
    floorPlan = None
    designator = data[2].value
    if data[0] != None and data[1] != None:
      if data[0].value != None: floorPlan = data[0].value
      else: floorplan = data[1].value
    if floorPlan and designator:
      if designator.upper().replace(' ','').isalpha() == False:
        temp = ' '.join(re.findall(r'\d+', designator.lower())).replace(' ','')
        position = int(''.join(list(map(str, temp))))
        if floorPlan in self.dict_floorPlans.keys():
          self.dict_floorPlans[floorPlan].append((designator,data[1].row, position))
        else: self.dict_floorPlans[floorPlan] = [(designator,data[1].row, position)]   
    return None
    
  def validate_floor_plans(self):
    #this function checks the occurrences of the floor plans
    #to ensure that it is not a typo or bad floor plan
    #typos logic is to be developed
    temp = {}
    adjusted = []
    name = []
    for k,v in self.dict_floorPlans.items():
      if k.lower().replace(" ","") in adjusted:
        temp[name[adjusted.index(k.lower().replace(" ",""))]].extend(v)
      else:
        temp[k] = v
        adjusted.append(k.lower().replace(" ",""))
        name.append(k)

    self.dict_floorPlans = temp
    L,P,S,WOW,W = [],[],[],[],[]
    temp = self.dict_floorPlans
    
    for k in temp.keys():
      L,P,S,WOW,W = [],[],[],[],[]
      for i in temp[k]:
        if i[0].lower().startswith('l'): L.append(i)
        elif i[0].lower().startswith('p'): P.append(i)
        elif i[0].lower().startswith('s'): S.append(i)
        elif i[0].lower().startswith('wow'): WOW.append(i)
        elif i[0].lower().startswith('w'): W.append(i)
      L.sort(key=lambda x: x[2])
      P.sort(key=lambda x: x[2])
      S.sort(key=lambda x: x[2])
      WOW.sort(key=lambda x: x[2])
      W.sort(key=lambda x: x[2])
      self.dict_floorPlans[k] = L + P + S + WOW + W
      
    self.check_designators_sequences()
    return None

  def check_designators_sequences(self):
    #this function checks the designator sequences and calls duplicates
    for k in self.dict_floorPlans.keys():
      size = len(self.dict_floorPlans[k])
      L,P,S,WOW,W = [],[],[],[],[]
      for i in range(size-1):
        first = self.dict_floorPlans[k][i]
        second = self.dict_floorPlans[k][i+1]
        if first[0] == second[0]:
          self.ws.cell(row=first[1], column=self.issueColumn).value += f"/is duplicates with row {second[1]}" 
          self.ws.cell(row=second[1], column=self.issueColumn).value += f"/is duplicates with row {first[1]}"
          self.list_color_cells.append(self.ws.cell(row=first[1], column=self.column_index_map["Flr Pln D"]))
          self.list_color_cells.append(self.ws.cell(row=second[1], column=self.column_index_map["Flr Pln D"]))
        elif "l" in first[0].lower() and "l" in second[0].lower():
          if second[2] != first[2] + 1:
            self.ws.cell(row=first[1], column=self.issueColumn).value += f"/skips a designator" 
            self.list_color_cells.append(self.ws.cell(row=first[1], column=self.column_index_map["Flr Pln D"]))
        elif "p" in first[0].lower() and "p" in second[0].lower():
          if second[2] != first[2] + 1: 
            self.ws.cell(row=first[1], column=self.issueColumn).value += f"/skips a designator"
            self.list_color_cells.append(self.ws.cell(row=first[1], column=self.column_index_map["Flr Pln D"]))
        elif "s" in first[0].lower() and "s" in second[0].lower():
          if second[2] != first[2] + 1: 
            self.ws.cell(row=first[1], column=self.issueColumn).value += f"/skips a designator"
            self.list_color_cells.append(self.ws.cell(row=first[1], column=self.column_index_map["Flr Pln D"]))
        elif first[0].lower().startswith("wow") and second[0].lower().startswith("wow"):
          if second[2] != first[2] + 1: 
            self.ws.cell(row=first[1], column=self.issueColumn).value += f"/skips a designator"
            self.list_color_cells.append(self.ws.cell(row=first[1], column=self.column_index_map["Flr Pln D"]))
        elif "w" in first[0].lower() and "wo" not in first[0].lower() and "w" in second[0].lower() and "wo" not in second[0].lower():
          if second[2] != first[2] + 1: 
            self.ws.cell(row=first[1], column=self.issueColumn).value += f"/skips a designator"
            self.list_color_cells.append(self.ws.cell(row=first[1], column=self.column_index_map["Flr Pln D"]))
      #print(f"\nfor {k}: {L},{P},{S},{W},{WOW}")
      print(f"\n{k}: {self.dict_floorPlans[k]}")
          
    return None

  def get_valid_floor_names(self, new_valid, old_valid):
    # Make a combined list of valid floor names:

    return None

  def get_invalid_floor_names(self, invalid_floors):
    #
    return None
# -------------------------------------------------
#end of set up function--------------------------------------------------------
  # -------------------------------------------------
# -------------------------------------------------
#start Processing functions----------------------------------------------------------
# -------------------------------------------------          
  def find_domain_issues(self, row_type_domain):
    # data comprised of the following:
    # data[0] = row number
    # data[1] = type
    # data[2] = IP
    # data[3] = Queue
    # data[4] = Make
    # data[5] = Model
    issue_message = ""
    # Ignore printers since they do not have domain field
    if "Printer" not in self.ws.cell(row=designator_type_domain[0], column=self.column_index_map["Type"]).value:
      if designator_type_domain[3] is not None and designator_type_domain[3] != "NSLIJHS.NET":
        issue_message += "/domain is not standard"
      elif designator_type_domain[3] is None and row_des_type_domain[4] != "":
        issue_message += "/domain is not inputted"
    # append issue message to the issues column
    self.ws.cell(row=row_dess_type_domain[0], column=self.issueColumn).value += issue_message

  def find_missing_floor_names(self, data):
    pass
    #this function checks for missing floor plans
    #floor data[1] and data[2] are the old and new floor plans respectively
    #they are located at K and L on excel sheet

    if data[1] is None and data[2] is None:
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln L"]))
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln N"]))
      return "/missing a floor plan"
    return ""

  def designatorIssues(self, data):
    #this function checks for missing designators and make sure it matches device type
    #they are located at M on the excel sheet
    issue_message = ""
    if data[1] is None:
      self.list_color_cells.append(
          self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln D"]))
      issue_message += "/missing a designator"
    elif data[1].lower()[0] not in "wl" and data[2] == "Workstation":
      issue_message += "designator doesn't match device type"
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln D"]))
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Type"]))
    elif data[1].lower()[0] not in "ps" and data[2] == "Printer":
      issue_message += "designator doesn't match device type"
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln D"]))
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Type"]))
    return issue_message

  def find_monitor_issues(self, data):
    #this function checks for missing monitors
    issue_message = ""
    if data[1] is not None or data[2] is not None or data[3] > 0:
      if data[1] is None:
        issue_message += "/1st monitor needs a make"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Make_1"]))
      if data[2] is None:
        issue_message += "/1st monitor needs a model"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Mod_1"]))
      if data[3] <= 0:
        issue_message += "/1st monitor needs a size"
        self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Mon 1"]))
    if data[4] is not None or data[5] is not None or data[6] > 0:
      if data[4] is None:
        issue_message += "/2nd monitor needs a make"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Make_2"]))
      if data[5] is None:
        issue_message += "/2nd monitor needs a model"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Mod_2"]))
      if data[6] <= 0:
        issue_message += "/2nd monitor needs a size"
        self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Mon 2"]))
    if data[7] is not None or data[8] is not None or data[9] > 0:
      if data[7] is None:
        issue_message += "/3rd monitor needs a make"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Make_3"]))
      if data[8] is None:
        issue_message += "/3rd monitor needs a model"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Mod_3"]))
      if data[9] <= 0:
        issue_message += "/3rd monitor needs a size"
        self.list_color_cells.append(
            self.ws.cell(row=data[0], column=self.column_index_map["Mon 3"]))
    if data[10] is not None or data[11] is not None or data[12] > 0:
      if data[10] is None:
        issue_message += "/4th monitor needs a make"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Make_4"]))
      if data[11] is None:
        issue_message += "/4th monitor needs a model"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Mod_4"]))
      if data[12] <= 0:
        issue_message += "/4th monitor needs a size"
        self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Mon 4"]))
    if issue_message:
      return issue_message
    return None

  def printer_Issues(self, data):
    # This function looks for printer errors
    # data1 = bc data2 = bf data3 = bd data4 = bg data5 = bh
    # bc = type bf = ip bd = queue name bg = make bh = model
    # data comprised of the following:
    # data[0] = row number
    # data[1] = type
    # data[2] = IP
    # data[3] = Queue
    # data[4] = Make
    # data[5] = Model
    issue_message = ""
    if data[1] is None:
      issue_message += "/Missing printer Type"
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["PRNT_Type"]))
    if data[2] is None:
      issue_message += "/Missing printer IP"
      self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["Network Pntr IP"]))
    if data[3] is None:
      issue_message += "/Missing printer Queue Name"
      self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["PRNT_Queue_Name"]))
    if data[4] is None:
      issue_message += "/Missing printer Make"
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["PRNT_Make"]))
    if data[5] is None:
      issue_message += "/Missing printer Model"
      self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["PRNT_Model"]))
  
    if issue_message:
      return issue_message
    return None

  def type_sequence(self, designator_type_domain):
    #this function checks to make sure the device type matches its designator
    #it also checks to make sure the domain is correct
    # parameters of the instance are row, designator, device type, domain
    issue_message = ""
    if designator_type_domain[1] is not None and designator_type_domain[2] is not None:
      if designator_type_domain[1].lower().startswith("wow"):
        pass
      elif designator_type_domain[1].lower().startswith('l'):
        if designator_type_domain[2].lower() not in "laptop":
          issue_message += "/device type doesn't match"
          self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0],column=self.column_index_map["Flr Pln D"]))
          self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0],column=self.column_index_map["WS_Type"]))
      elif designator_type_domain[1].lower().startswith('w'):
        if designator_type_domain[2].lower() not in "desktop":
          issue_message += "/device type doesn't match"
          self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0],column=self.column_index_map["Flr Pln D"]))
          self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0], column=self.column_index_map["WS_Type"]))
          self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0],column=self.column_index_map["Domain"]))
    return issue_message

  def check_duplicate(self, data):
    #this function checks if cell is duplicated and adds it to list if it is
    return None
#end of issue functions----------------------------------------------------------
#####################################################
######################################################
######################################################
######################################################
#implementation functions------------------------------------------------------------

  def flagging_issues(self):
    #this caller function will loop each row and call issue functions
    issue_messages = []
    for i in range(3, self.ws.max_row + 1):

      #Collect all issue messages for the current row
      row_issues = []

      #Check for floorplan issues
      row_issues.append(self.find_missing_floor_names([
              i,self.ws.cell(row=i,column=self.column_index_map["Flr Pln N"]).value,
              self.ws.cell(row=i,column=self.column_index_map["Flr Pln L"]).value
          ]))

      #Check for designator issues
      row_issues.append(self.designatorIssues([
              i,self.ws.cell(row=i,column=self.column_index_map["Flr Pln D"]).value,
              self.ws.cell(row=i, column=self.column_index_map["Type"]).value
          ]))

      #Check for device type issues
      row_issues.append(self.type_sequence([
              i,
              self.ws.cell(row=i,column=self.column_index_map["Flr Pln D"]).value,
              self.ws.cell(row=i,column=self.column_index_map["WS_Type"]).value,
              self.ws.cell(row=i, column=self.column_index_map["Domain"])
          ]))

      #Check for monitor issues
      if "workstation" in self.ws.cell(
          row=i, column=self.column_index_map["Type"]).value.lower():
        params = [
            i,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Make_1"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Mod_1"]).value,
            self.ws.cell(row=i, column=self.column_index_map["Mon 1"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Make_2"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Mod_2"]).value,
            self.ws.cell(row=i, column=self.column_index_map["Mon 2"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Make_3"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Mod_3"]).value,
            self.ws.cell(row=i, column=self.column_index_map["Mon 3"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Make_4"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Mod_4"]).value,
            self.ws.cell(row=i, column=self.column_index_map["Mon 4"]).value
        ]
        row_issues.append(self.find_monitor_issues(params))

      #Check for printer issues
      if "printer" in self.ws.cell(
          row=i, column=self.column_index_map["Type"]).value.lower():
        row_issues.append(
            self.printer_Issues([
                i,
                self.ws.cell(row=i,column=self.column_index_map["PRNT_Type"]).value,
                self.ws.cell(row=i,column=self.column_index_map["Network Pntr IP"]).value,
                self.ws.cell(row=i,column=self.column_index_map["PRNT_Queue_Name"]).value,
                self.ws.cell(row=i,column=self.column_index_map["PRNT_Make"]).value,
                self.ws.cell(row=i,column=self.column_index_map["PRNT_Model"]).value
            ]))

      #check for duplicates

      
      #Collect issue messages for the current row
      issue_messages.append(row_issues)

    #output devices to the row
    for i in range(3, self.ws.max_row + 1):
      combined_message = " ".join(msg for msg in issue_messages[i - 3] if msg)
      self.ws.cell(row=i, column=self.issueColumn).value += combined_message

    return None
    
  def process_data(self):
    print(f"Beginning of process_data: {self.output_file}")
    self.setup_logging()
    self.load_Data()
    self.flagging_issues()
    self.validate_floor_plans()
    self.highlight_rows()
    self.save_output_file()
    print(f"Ending of process_data: {self.output_file}")
    return None


# -------------------------------------------------
# End of Processing Functions
# -------------------------------------------------

# -------------------------------------------------
# Start Implementation Functions------------------------------------------------------------
# -------------------------------------------------

# Instantiate the DataProcessor class
print("Starting the data the processor")
data_processor = DataProcessor()

# Call the process_data method on the instance
data_processor.process_data()

print("After calling the process_data method on the instance")
# Configure logging to write messages/logs to a file
logging.basicConfig(filename='data_processing.log', level=print)

# -------------------------------------------------
# End of Implementation Functions
# -------------------------------------------------
