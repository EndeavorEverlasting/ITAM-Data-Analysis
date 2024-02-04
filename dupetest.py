import re, os, sys, logging, random
from shutil import copy2
#from typing import Union
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


class DataProcessor:

  def update_headers(self):
    # Remove the "Department_ID" column
    for row in self.ws.iter_rows(min_row=1, max_row=1):
      for cell in row:
        if cell.value == "Department_ID":
          self.ws.delete_cols(cell.column, 1)
          break

    # Update other headers and column index map
    header_row = self.ws[1]
    column_names = [cell.value for cell in header_row]
    self.column_index_map = {
        name: index + 1
        for index, name in enumerate(column_names)
    }

  def __init__(self):
    self.input_file = "LIJ 2_2_24.xlsx"
    self.output_file = None

  def load_Data(self):
    # This function loads data from the excel sheet and adds a column for issues
    # The "Issues" column will be the first column (column 1)

    self.wb = load_workbook(self.input_file, read_only=False, data_only=False)
    self.ws = self.wb.active  # This part picks the first sheet on the excel

    # Call update_headers() to remove "Department_ID" and update other headers
    self.update_headers()

    # Add the "Issues" column at the same column position as "Department_ID"
    self.issueColumn = 6  # Column F
    self.ws.cell(row=2, column=self.issueColumn).value = "Issues"

    self.initialize_column_index_map()  # Run after adding column for issues
    for c in self.ws["CD"]:
      c.value = ""
    return None

  def add_issueColumn(self):
    # Shift existing columns to the right to make space for the "Issues" column
    for column in range(self.ws.max_column, 0, -1):
      for row in range(3, self.ws.max_row + 1):
        self.ws.cell(row=row, column=column + 1).value = self.ws.cell(
            row=row, column=column).value
      self.ws.cell(row=row, column=column).value

    # Add the "Issues" column at the first column position in row 2
    self.ws.cell(row=2, column=1).value = "Issues"

  def initialize_column_index_map(self):
    # This function initializes a dictionary to map column names to column indices
    header_row = self.ws[2]
    column_names = [cell.value for cell in header_row]
    self.column_index_map = {
        name: index + 1
        for index, name in enumerate(column_names)
    }

    print("Header Row: ", header_row)
    print("Column Index Map:", self.column_index_map)

  def floorPlanIssues(self, data):
    #this function checks for missing floor plans
    #floor data[1] and data[2] are the old and new floor plans respectively
    #they are located at K and L on excel sheet
    
    if data[1] == None and data[2] == None:
      self.ws[f"CD{data[0]}"].value += "/needs a floor plan"
    return None

  def designatorIssues(self, data):
    #this function checks for missing designators
    #they are located at M on the excel sheet
    if data[1] is None:
      return "/needs a designator"
    return ""

  def monitorIssues(self, data):
    #this function checks for missing monitors
    #3 things can be missing on monitor info at Y,Z,AA
    if self.ws[f"B{data[0]}"].value.lower() in "workstation,laptop,desktop":
      if data[1] == None:
        self.ws.cell(row=data[0], column=self.issueColumn).value = str(
            self.ws.cell(
                row=data[0],
                column=self.issueColumn).value) + "/needs a monitor make"
      if data[2] == None:
        self.ws.cell(row=data[0], column=self.issueColumn).value = str(
            self.ws.cell(
                row=data[0],
                column=self.issueColumn).value) + "/needs a monitor model"
      if data[3] == None:
        self.ws.cell(row=data[0], column=self.issueColumn).value = str(
            self.ws.cell(
                row=data[0],
                column=self.issueColumn).value) + "/needs a monitor size"
    return None

  def check_duplicate_designators(self):
    duplicates = {}
    for i in range(3, self.ws.max_row + 1):
      # Extract the shared portion of the floor plan name
      floor_plan_shared = re.search(
          r'^LIJMC - \d+',
          self.ws.cell(row=i, column=self.column_index_map["Flr Pln N"]).value)
      if floor_plan_shared:
        floor_plan_shared = floor_plan_shared.group()
      else:
        floor_plan_shared = ""

      condition = (
          self.ws.cell(row=i, column=self.column_index_map["Flr Pln L"]).value,
          floor_plan_shared,  # Use the shared portion for comparison
          self.ws.cell(row=i, column=self.column_index_map["Flr Pln D"]).value)

      if condition in duplicates:
        duplicates[condition].append(i)
      else:
        duplicates[condition] = [i]

    for condition, rows in duplicates.items():
      if len(rows) > 1:
        self.ws.cell(row=rows[0], column=self.issueColumn).value = (
            f"Duplicates with {', '.join(str(row) for row in rows if row != rows[0])}"
        )

  def sequence_check(self, data):
    try:
      floorPlan = None
      if data[-1] is not None:
        if data[1] is None and data[2] is not None:
          floorPlan = data[2][-1]
        elif data[1] is not None and data[2] is None:
          floorPlan = data[1][-1]
        if floorPlan and len(floorPlan) < 10:
          designator = str(data[-1])
          if isinstance(designator, str) and len(designator) > 1:
            if designator.lower().startswith("wow"):
              if designator[3] != floorPlan:
                self.ws.cell(row=data[0], column=self.issueColumn).value += (
                    "/floor plan and designator are different")
              else:
                self.ws.cell(row=data[0], column=self.issueColumn).value += (
                    "/floor plan and designator are correct")
          else:
            if designator[1] != floorPlan:
              self.ws.cell(row=data[0], column=self.issueColumn).value += (
                  "/floor plan and designator are different")
            else:
              self.ws.cell(row=data[0], column=self.issueColumn).value += (
                  "/floor plan and designator are correct")
    except Exception as e:
      print(
          f"Error in sequence_check: {e}\nLine: {sys.exc_info()[-1].tb_lineno}"
      )
    return None

  def printer_Issues(self, data):
    # This function looks for printer errors
    # data1 = bc data2 = bf data3 = bd data4 = bg data5 = bh
    # bc = type bf = ip bd = queue name bg = make bh = model
    issue_message = ""
    if data[1] is None:
      issue_message += "/Missing printer Type"
    if data[2] is None:
      issue_message += "/Missing printer IP"
    if data[3] is None:
      issue_message += "/Missing printer Queue Name"
    if data[4] is None:
      issue_message += "/Missing printer Make"
    if data[5] is None:
      issue_message += "/Missing printer Model"

    if issue_message:
      current_issue_value = str(
          self.ws.cell(row=data[0], column=self.issueColumn).value)
      if current_issue_value:
        self.ws.cell(row=data[0],
                     column=self.issueColumn).value += issue_message
      else:
        self.ws.cell(row=data[0],
                     column=self.issueColumn).value = issue_message
    return None

  def flaggingIssues(self):
    for i in range(3, self.ws.max_row + 1):
      # Collect issue messages
      issue_messages = []
      issue_messages.append(self.floorPlanIssues([i, self.ws[f"K{i}"].value, self.ws[f"L{i}"].value]))
      issue_messages.append(self.designatorIssues([i, self.ws[f"M{i}"].value]))
      # Collect issue messages from other functions as well

      # Combine issue messages into one string
      combined_message = " ".join(msg for msg in issue_messages if msg)

      # Update the "Issues" column
      self.ws.cell(row=i, column=self.issueColumn).value = combined_message
    
    #checks for issues with monitor designator and floorplans
    for i in range(3, self.ws.max_row + 1):
      self.floorPlanIssues([i, self.ws[f"K{i}"].value, self.ws[f"L{i}"].value])
      self.designatorIssues([i, self.ws[f"M{i}"].value])
      #list may need to extend in future because monitor goes up to 4
      self.monitorIssues([
          i, self.ws[f"Y{i}"].value, self.ws[f"Z{i}"].value,
          self.ws[f"AA{i}"].value
      ])
      # i is the row-index of ws.self.ws[f"A{i}"].value
      # self.ws[f"K{i}"].value is the Flr Pln L in the cell K[i]
      # self.ws[f"L{i}"].value is the Flr Pln N in the cell L[i]
      # self.ws[f"M{i}"].value is the Flr Pln D in the cell M[i]
      self.sequence_check([
          i,
          self.ws.cell(row=i, column=self.column_index_map["Flr Pln L"]).value,
          self.ws.cell(row=i, column=self.column_index_map["Flr Pln N"]).value,
          self.ws.cell(row=i, column=self.column_index_map["Flr Pln D"]).value
      ])
      if "printer" in self.ws.cell(
          row=i, column=self.column_index_map["Type"]).value.lower():
        self.printer_Issues([
            i,
            self.ws.cell(row=i,
                         column=self.column_index_map["PRNT_Type"]).value,
            self.ws.cell(
                row=i, column=self.column_index_map["Network Pntr IP"]).value,
            self.ws.cell(
                row=i, column=self.column_index_map["PRNT_Queue_Name"]).value,
            self.ws.cell(row=i,
                         column=self.column_index_map["PRNT_Make"]).value,
            self.ws.cell(row=i,
                         column=self.column_index_map["PRNT_Model"]).value
        ])
    self.check_duplicate_designators()
    return None

  def highlight_duplicates(self):
    try:
      for i in range(3, self.ws.max_row + 1):
        duplicate_message = self.ws.cell(row=i, column=self.issueColumn).value
        if "Duplicates with" in duplicate_message:
          # Extract IDs from the message generated in check_duplicate_designators()
          ids = re.findall(r'\d+', duplicate_message)
          if len(ids) == 2:
            new_id, existing_id = map(int, ids)
            # Apply red fill to the cell
            for col_index in range(2, self.ws.max_column + 1):
              cell = self.ws.cell(row=new_id,
                                  column=col_index)  # Use new_id instead of i
              cell.fill = PatternFill(start_color='FF0000',
                                      end_color='FF0000',
                                      fill_type='solid')
          else:
            logging.warning("Issue with the number of IDs in duplicates.")
    except Exception as e:
      print(f"Error occurred while highlighting duplicates: {e}")

  def highlight_sequence_errors(self):
    try:
      for i in range(3, self.ws.max_row + 1):
        # Your sequence error logic here
        pass
    except Exception as e:
      print(f"Error occurred while highlighting sequence errors: {e}")

  def highlight_printer_issues(self):
    try:
      for i in range(3, self.ws.max_row + 1):
        # Your printer issues logic here
        pass
    except Exception as e:
      print(f"Error occurred while highlighting printer issues: {e}")

  def highlight_other_issues(self):
    try:
      for row in range(3, self.ws.max_row + 1):
        issue_value = self.ws.cell(row=row, column=self.issueColumn).value
        if issue_value:
          color = None
          if len(str(issue_value).split('/')) == 2:
            color = "FFFF00"  # Yellow
          elif len(str(issue_value).split('/')) == 3:
            color = "FF8000"  # Orange
          elif len(str(issue_value).split('/')) > 3:
            color = "FF0000"  # Red
          if color:
            # Apply fill color to the entire row
            for col_index in range(1, self.ws.max_column + 1):
              self.ws.cell(row=row, column=col_index).fill = PatternFill(
                  start_color=color, end_color=color, fill_type="solid")
    except Exception as e:
      print(f"Error occurred while highlighting other issues: {e}")

  def highlight_Issues(self):
    # Highlight specific types of issues
    self.highlight_duplicates()
    self.highlight_sequence_errors()
    self.highlight_printer_issues()

    # Highlight other issues based on frequency of errors
    self.highlight_other_issues()

  def highlight_duplicate_des(self, duplicate_message=None):
    logging.info(
        f"Beginning of highlight_duplicate_des(self, duplicates): {self.output_file}"
    )
    try:
      # Check if workbook and sheetnames exist.
      if self.wb is not None and self.wb.sheetnames:
        sheet_name = self.wb.sheetnames[0]
        sheet = self.wb[sheet_name]
        new_id = None
        existing_id = None

        # Extract IDs from the message generated in check_duplicate_designators()
        ids = re.findall(r'\d+', duplicate_message)
        if len(ids) == 2:
          new_id, existing_id = map(int, ids)

          # Find the row index for the new_id
          new_id_row = None
          for row_index in range(2, sheet.max_row + 1):
            current_id_cell = sheet.cell(row=row_index, column=1)
            current_id_value = current_id_cell.value
            if current_id_value == new_id:

              try:
                # Attempt to convert the current ID to an integer
                current_id = int(current_id_value)
              except (ValueError, TypeError):
                # Handle cases where 'ID' value is not a valid integer
                current_id = None
              if current_id == new_id:
                new_id_row = row_index
                break

          #Check if the new_id_row is found
          if new_id_row is not None:
            # Apply red fill to the cell
            for col_index in range(2, sheet.max_column + 1):
              cell = sheet.cell(row=new_id_row, column=col_index)
              cell.fill = PatternFill(start_color='FF0000',
                                      end_color='FF0000',
                                      fill_type='solid')

            # Save the workbook
            self.wb.save(filename=self.output_file)
            print(f"File Has Been Updated in {self.output_file}")

        else:
          logging.warning("Issue with the number of IDs.")

      else:
        logging.warning("Workbook is not defined. Unable to save.")

    except Exception as e:
      print(f"Error occurred while highlighting duplicates: {e}")

    finally:
      logging.info(
          f"Ending of highlight_duplicate_des(self, duplicate_message): "
          f"{self.output_file}")

  def save_output_file(self):
    logging.info(f"Beginning the determine_output_files(): {self.output_file}")
    output_folder = 'Outputs/'
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_extension = os.path.splitext(self.input_file)[1].lower()
    file_type = {
        'csv': 'csv',
        'xlsx': 'ss',
        'txt': 'txt',
        'log': 'logs'
    }.get(file_extension, 'ss')

    output_file_name = f"output_{current_datetime}.xlsx"
    self.output_file = os.path.join(output_folder, file_type, output_file_name)

    # Check if the output_file already exists
    if os.path.exists(self.output_file):
      # Add a copy number if the file already exists
      copy_number = 1
      while os.path.exists(self.output_file):
        output_file_name = f"output_{current_datetime}_copy{copy_number}.xlsx"
        self.output_file = os.path.join(output_folder, file_type,
                                        output_file_name)
        copy_number += 1

    self.wb.save(self.output_file)
    logging.info(f"Ending the save_output_file(): {self.output_file}")
    print(self.output_file)
    return None

  def setup_Logging(self):
    #this function sets up logging, and sets up the directory to log data to.
    output_folder = 'Outputs/'
    output_directory = os.path.join(output_folder, 'logs')
    os.makedirs(output_directory, exist_ok=True)
    log_file = os.path.join('Outputs', 'logs', 'data_processor.log')
    logging.basicConfig(
        filename=log_file,  # Set the log file name
        level=logging.INFO,  # Set the logging level to INFO
        format='%(asctime)s - %(levelname)s - %(message)s')  # format 3

  def process_data(self):
    logging.info(f"Beginning of process_data: {self.output_file}")
    self.setup_Logging()
    self.load_Data()
    self.flaggingIssues()
    self.highlight_Issues()
    self.highlight_duplicate_des(
        "dummy message")  # Call the method with a dummy message
    self.save_output_file()
    logging.info(f"Ending of process_data: {self.output_file}")
    return None


# Run the current data procesor
data_processor = DataProcessor().process_data()

# Configure logging to write messages/logs to a file
logging.basicConfig(filename='data_processing.log', level=logging.INFO)
