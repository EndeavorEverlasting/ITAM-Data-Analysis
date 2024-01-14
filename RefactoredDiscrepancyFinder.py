import os
import logging
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

incorrect_designators = []
missing_floorplans = []
duplicates = []
missing_makes = []
missing_models = []
missing_sizes = []

# Configure logging to write messages to a file
# Console output is limited on replit
# logging.basicConfig(filename='log.txt', level=logging.INFO)


def highlight_duplicates(df, output_file):
  try:
    book = load_workbook(output_file)
    print(f"Reading the Excel file: {output_file}")
  except FileNotFoundError:
    # If file doesn't exist, create a new workbook
    book = load_workbook()
    book.save(output_file)
    print(f"Created a new Excel file: {output_file}")

    sheet_name = "Sheet1"
    sheet = book[sheet_name]

    for duplicate_id in duplicates:
      # Adjust for 1-based index and header row
      row_index = int(duplicate_id.split()[0]) + 2
      for col_index in range(1, df.shape[1] + 1):
        cell = sheet.cell(row=row_index, column=col_index)
        cell.fill = PatternFill(start_color='FFFF0000',
                                end_color='FFFF0000',
                                fill_type='solid')


#      df.at[int(duplicate_id.split()[0]),
#           'ID'].style.applymap(lambda x: 'background-color: red',
#                               subset=['ID']).to_excel(writer,
#                                                      sheet_name,
#                                                     index=False,
#                                                    startrow=1,
#                                                   header=False)

# Save the changes to an Excel file
      book.save(output_file)
      print(f"File Has Been Updated in {output_file}")

  return None


def findProblems(input_file, output_file):
  # Read the Excel file
  df = pd.read_excel(input_file, skiprows=1, engine='openpyxl')

  # Clean up 'Flr Pln D' column by removing leading and trailing spaces
  df['Flr Pln D'] = df['Flr Pln D'].str.strip()
  #this line iterates over each row to get its single data for that row
  for index, row in df.iterrows():
    # print(f"Processing row {index} of {len(df)}")
    designator = row['Flr Pln D']
    epic_loc = row['EPIC_LOC']
    department = row['Department']
    newFloorplan = row['Flr Pln N']
    oldFloorplan = row['Flr Pln L']

    monitor = monitorIssues(row)
    # designator = designatorIssues(row)
    floorplan = floorPlanIssues(row)
    #print(flaggingIssues(monitor,designator,floorplan,row))

  checkDuplicateDesignators(df)

  print(duplicates)
  print(f'Incorrect Designators: {len(incorrect_designators)}')
  print(f'Duplicates: {len(duplicates)}')
  print(f'Missing Data: {len(missing_floorplans)}')
  print(f'Monitor Makes Missing: {len(missing_makes)}')
  #Add cases for models and sizes
  #Highlight discrepancies for users:
  highlight_duplicates(
      df, output_file)  # Only gets called when findProblems is called

  return None


def designatorIssues(data):
  response = ""
  global incorrect_designators
  des = data['Flr Pln D']
  if not pd.notna(des):
    incorrect_designators.append(data[des])
    response = " needs a designator"
  return response


def checkDesignatorSequence(data):
  response = ""
  global incorrect_designators
  # appends IDs to incorrect designators to a list
  return None


def floorPlanIssues(data):
  response = ""
  global missing_floorplans
  plan_n = data['Flr Pln N']
  plan_l = data['Flr Pln L']
  if (not pd.notna(plan_n)) and (not pd.notna(plan_l)):
    missing_floorplans.append(data["ID"])
    response = " needs a floorplan"
  return response


def checkDuplicateDesignators(data):
  '''
  need to cross check floor plans L,N,D if floor plan 
  and designator is the same then its a duplicate.
  it is done by looping over the the data and 
  if it doesnt match the criteria then it adds it to first IDS.
  then add it's data to checklis. 
  if it matches duplicate conditions then it adds it to IDs_of_Duplicates and
  adds the condition to finder. 
  then loop over finder to find the same condition in checklis and get its index.
  with the index we can match the duplicated id with the first found it.
  '''
  firstIds = []
  checklis = []
  IDs_of_Duplicates = []
  finder = []
  global duplicates
  for i, r in data.iterrows():
    # Sometimes, a user will mark a record as "NOT FOUND" in Flr Pln D
    # Other times, a user with mark a device as remote and will not input a designator
    # Skip such cases in search for duplicates:
    if str(r['Flr Pln D']).lower() or str(
        r['EPIC_LOC']).lower() not in "not found":
      condition_1 = f"{r['Flr Pln D']}/{r['Flr Pln N']}"
      condition_2 = f"{r['Flr Pln D']}/{r['Flr Pln L']}"
      condition_3 = f"{r['Flr Pln D']}/{r['Department']}"
      if condition_1 not in str(checklis) or condition_2 not in str(
          checklis) or condition_3 not in str(checklis):
        firstIds.append(str(r['ID']))
        checklis.append(condition_1 + condition_2 + condition_3)
      else:
        IDs_of_Duplicates.append(str(r['ID']))
        finder.append(condition_1 + condition_2 + condition_3)

  for f in finder:
    if f in checklis:
      n = checklis.index(f)
      # This line prints the index of the condition f in the checklis array.
      # print(checklis.index(f))
      duplicates.append(
          f"{firstIds[n]} is duplicates with {IDs_of_Duplicates[finder.index(f)]}"
      )
  #print(firstIds)
  #print(checklis)
  #print(IDs_of_Duplicates)
  return None


def monitorIssues(data):
  #missing monitors or the monitor's wrong size what ever issues with it
  response = ""
  global missing_makes
  global missing_models
  global missing_sizes
  # The following block is suscpetible to errors
  # because of the way the data is formatted.
  # There might be a blank or an empty string in the data.
  # We'll try and catch the logic to handle the error acutely.
  try:
    # print(f"Row Type: {str(data['Type']).lower()}")
    if str(data['Type']).lower() in "workstation,laptop,desktop":
      # print(f"Values in data: {data.values.tolist()}")
      make = data['WS_Mon_Make_1']
      model = data['WS_Mon_Mod_1']
      size = data['Mon 1']
      if not pd.notna(make):
        missing_makes.append(data["ID"])
        response = "needs a monitor make"
      elif not pd.notna(model):
        missing_models.append(data["ID"])
        response = "needs a monitor model"
      # size is a number stored as a string so
      # we need to typecast it to an integer and check if it equals zero
      elif not pd.notna(size) or int(size) == 0:
        missing_sizes.append(data["ID"])
        response = "needs a monitor size"
      else:
        response = "Unknown 4th Issue"
  except KeyError as e:
    print(f"KeyError: {e}, Row Data: {data.values.tolist()}")
  return response


def flaggingIssues(mon, desi, floor, row):
  #adds the issue from the called function to the data and file it
  return f"{row['ID']} {mon} {desi} {floor}"


def outputToTextFile():
  #this function will output the data to a text file
  return None


def outputToCSVFile():
  #this function will output the data to a CSV file
  return None


def outputToXLSXFile():
  #this function will output the data to an xlsx file
  return None


def searchForIncorrectType():
  #this functions will compare host name and designator with type
  #if the two determine that type is wrong or vice-versa, it will flag
  return None


input_file = 'lij 1.12.24.xlsx'
output_file = 'template.xlsx'
findProblems(input_file, output_file)
# highlight_duplicates(pd.DataFrame(), output_file)
