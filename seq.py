import contextlib
import logging
import os

#from typing import Union
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class DataProcessor:

  def __init__(self):
    self.input_file = "LIJ 2_2_24.xlsx"
    self.output_file = None
    self.list_color_cells = []
    self.list_duplicates = []
    self.list_valid_floorPlans = []

# -------------------------------------------------
#start set up functions-------------------------------------------------
# -------------------------------------------------

  import os
  from datetime import datetime

  def save_output_file(self):
      print(f"Beginning the determine_output_files(): {self.output_file}")
      output_folder = 'Outputs/'
      current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
      file_extension = os.path.splitext(self.input_file)[1].lower()
      file_type = {
          'csv': 'csv',
          'xlsx': 'ss',
          'txt': 'txt',
          'log': 'logs'
      }.get(file_extension, 'ss')

      output_file_name = f"output_{current_datetime}.xlsx"
      output_directory = os.path.join(output_folder, file_type)

      # Ensure the output directory exists
      os.makedirs(output_directory, exist_ok=True)  # This will create the directory if it doesn't exist

      self.output_file = os.path.join(output_directory, output_file_name)

      # Check if the output_file already exists
      if os.path.exists(self.output_file):
          # Add a copy number if the file already exists
          copy_number = 1
          while os.path.exists(self.output_file):
              output_file_name = f"output_{current_datetime}_copy{copy_number}.xlsx"
              self.output_file = os.path.join(output_directory, output_file_name)
              copy_number += 1

      self.wb.save(self.output_file)
      print(f"Ending the save_output_file(): {self.output_file}")
      print(self.output_file)
      return None

  def load_Data(self):
    # This function loads data from the excel sheet and adds a column for issues
    # The "Issues" column will be the first column (column 1)

    self.wb = load_workbook(self.input_file, read_only=False, data_only=False)
    self.ws = self.wb.active  # This part picks the first sheet on the excel

    # Call update_headers() to remove "Department_ID" and update other headers
    self.update_headers()

    # Add the "Issues" column at the same column position as "Department_ID"
    self.issueColumn = 6  # Column F
    self.ws.cell(row=2, column=self.issueColumn).value = "Issues"

    self.initialize_column_index_map()  # Run after adding column for issues
    for c in self.ws["CD"]:
      c.value = ""
    for x in range(3, self.ws.max_row + 1):
      self.ws.cell(row=x, column=self.issueColumn).value = ""
    return None

  def update_headers(self):
    header_row = self.ws[1]
    column_names = [cell.value for cell in header_row]
    self.column_index_map = {
        name: index + 1
        for index, name in enumerate(column_names)
    }

  def initialize_column_index_map(self):
    # This function initializes a dictionary to map column names to column indices
    header_row = self.ws[2]
    column_names = [cell.value for cell in header_row]
    self.column_index_map = {
        name: index + 1
        for index, name in enumerate(column_names)
    }

    print("Header Row: ", header_row)
    print("Column Index Map:", self.column_index_map)

  def setup_logging(self):
    #this function sets up logging, and sets up the directory to log data to.
    output_folder = 'Outputs/'
    output_directory = os.path.join(output_folder, 'logs')
    os.makedirs(output_directory, exist_ok=True)
    log_file = os.path.join('Outputs', 'logs', 'data_processor.log')
    logging.basicConfig(
        filename=log_file,  # Set the log file name
        level=logging.INFO,  # Set the logging level to INFO
        format='%(asctime)s - %(levelname)s - %(message)s')  # format 3

  def find_duplicates_by_floor_name(self, valid_floors):
    # This function finds duplicate designators in the excel sheet and adds them to a list
    # For each floor name, 
    pass

  def find_sequence_errors(self, valid_floors):
    # This function iterates over the floor names and will check designators for sequential errors
    pass

  def highlight_rows(self):
    #this function loops through the data and highlight the rows with issues
    #then it loops through the cell issues list and highlight those cells
    yellow = "00FFFF66"
    orange = "00FF9933"
    red = "00FF0000"
    pink = "00FF66FF"
    for i in range(3, self.ws.max_row + 1):
      col = self.ws.cell(row=i, column=self.issueColumn).value
      if len(col.split("/")) == 2:
        for n in range(1, 82):
          x = self.ws.cell(row=i, column=n)
          x.fill = PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
      elif len(col.split("/")) == 3:
        for n in range(1, 82):
          x = self.ws.cell(row=i, column=n)
          x.fill = PatternFill(start_color=orange,end_color=orange,fill_type="solid")
      elif len(col.split("/")) >= 4:
        for n in range(1, 82):
          x = self.ws.cell(row=i, column=n)
          x.fill = PatternFill(start_color=red,end_color=red,fill_type="solid")
      else:
        pass

    for i in self.list_color_cells:
      i.fill = PatternFill(start_color=pink, end_color=pink, fill_type="solid")
    return None

  def validate_floor_plans(self):
    valid_floor_plans = {"New": {}, "Old": {}}
    for i in range(3, self.ws.max_row + 1):
        for column_name in ["Flr Pln N"]:
            floor_plan = self.ws.cell(row=i, column=self.column_index_map[column_name]).value
            if floor_plan:
                # Determine if the floor plan is new or old based on its format
                category = "New" if "-" in floor_plan.split()[:2] else "Old"
                if floor_plan not in valid_floor_plans[category]:
                    valid_floor_plans[category][floor_plan] = 1
                else:
                    valid_floor_plans[category][floor_plan] += 1

    # Filter out floor plans not meeting the validation criteria (> 3 designators)
    for category in ["New", "Old"]:
        valid_floor_plans[category] = {k: v for k, v in valid_floor_plans[category].items() if v > 3}

    # Prepare a list of valid floor plans for other functions
    self.list_valid_floorPlans = [fp for category in valid_floor_plans.values() for fp in category.keys()]

    print(f"\n\nValid floor plan list: {self.list_valid_floorPlans}")
    
    #Send valid floor plans to the other functions
    self.count_devices(valid_floor_plans)
    # self.duplicates()
    # self.sequence_errors()

  def check_designators_duplicates(self, valid_floors):
    #this function checks for duplicates in the designators
    pass

  def check_designators_sequences(self, valid_floors):
    pass

  def get_floor_name_validity(self, invalid_floors, new_floors, old_floors):
    # Checks to see if the floor names are valid
    # by using the validate_floor_plans() function
    print(f"start of get_floor_name_validity(): {list(invalid_floors)}")
    is_valid_floor_name = False
    for k, v in new_floors.items():
      if k in invalid_floors:
        print(f"Invalid floor name: {k}")
        is_valid_floor_name = False
        break
      elif v in invalid_floors:
        is_valid_floor_name = False
        break
      else:
        is_valid_floor_name = True
        break
    for k, v in old_floors.items():
      if k in invalid_floors:
        print(f"Invalid floor name: {k}")
        is_valid_floor_name = False
        break
      elif v in invalid_floors:
        is_valid_floor_name = False
        break
      else:
        is_valid_floor_name = True
        break
    return is_valid_floor_name

  def get_valid_floor_names(self, new_valid, old_valid):
    # Make a combined list of valid floor names:
    combined_valid = {}
    # How to combine the two dictionaries?

    return combined_valid

  def get_invalid_floor_names(self, invalid_floors):
    # mark designators with the invalid floor names since names are already highlighted
    # this function will return the invalid floor names object
    print(type(invalid_floors))
    print(list(invalid_floors))
    return invalid_floors
# -------------------------------------------------
#end of set up function--------------------------------------------------------
  # -------------------------------------------------
# -------------------------------------------------
#start Processing functions----------------------------------------------------------
# -------------------------------------------------

  def is_valid_floor_plan(self, valid_floor_plans, floor_plan):
    # Tests a passed in floor plan to see if it is valid
    if (floor_plan in valid_floor_plans):
      print(f"floor plan {floor_plan} is valid")
      return True
    else:
      print(f"Invalid floor plan: {floor_plan}")
      # highlight floor plan cell a light green
      # pass data from cell with an issue into the highlight_cell() method
      return False

  def find_domain_issues(self, row_type_domain):
    # data comprised of the following:
    # data[0] = row number
    # data[1] = type
    # data[2] = IP
    # data[3] = Queue
    # data[4] = Make
    # data[5] = Model
    issue_message = ""
    # Ignore printers since they do not have domain field
    if "Printer" not in self.ws.cell(row=designator_type_domain[0], column=self.column_index_map["Type"]).value:
      if designator_type_domain[3] is not None and designator_type_domain[3] != "NSLIJHS.NET":
        issue_message += "/domain is not standard"
      elif designator_type_domain[3] is None and row_des_type_domain[4] != "":
        issue_message += "/domain is not inputted"
    # append issue message to the issues column
    self.ws.cell(row=row_dess_type_domain[0], column=self.issueColumn).value += issue_message

  def find_missing_floor_names(self, data):
    #this function checks for missing floor plans
    #floor data[1] and data[2] are the old and new floor plans respectively
    #they are located at K and L on excel sheet

    if data[1] is None and data[2] is None:
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln L"]))
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln N"]))
      return "/missing a floor plan"
    return ""

  def designatorIssues(self, data):
    #this function checks for missing designators and make sure it matches device type
    #they are located at M on the excel sheet
    issue_message = ""
    if data[1] is None:
      self.list_color_cells.append(
          self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln D"]))
      issue_message += "/missing a designator"
    elif data[1].lower()[0] not in "wl" and data[2] == "Workstation":
      issue_message += "designator doesn't match device type"
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln D"]))
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Type"]))
    elif data[1].lower()[0] not in "ps" and data[2] == "Printer":
      issue_message += "designator doesn't match device type"
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Flr Pln D"]))
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Type"]))
    return issue_message

  def find_monitor_issues(self, data):
    #this function checks for missing monitors
    issue_message = ""
    if data[1] is not None or data[2] is not None or data[3] > 0:
      if data[1] is None:
        issue_message += "/1st monitor needs a make"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Make_1"]))
      if data[2] is None:
        issue_message += "/1st monitor needs a model"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Mod_1"]))
      if data[3] <= 0:
        issue_message += "/1st monitor needs a size"
        self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Mon 1"]))
    if data[4] is not None or data[5] is not None or data[6] > 0:
      if data[4] is None:
        issue_message += "/2nd monitor needs a make"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Make_2"]))
      if data[5] is None:
        issue_message += "/2nd monitor needs a model"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Mod_2"]))
      if data[6] <= 0:
        issue_message += "/2nd monitor needs a size"
        self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Mon 2"]))
    if data[7] is not None or data[8] is not None or data[9] > 0:
      if data[7] is None:
        issue_message += "/3rd monitor needs a make"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Make_3"]))
      if data[8] is None:
        issue_message += "/3rd monitor needs a model"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Mod_3"]))
      if data[9] <= 0:
        issue_message += "/3rd monitor needs a size"
        self.list_color_cells.append(
            self.ws.cell(row=data[0], column=self.column_index_map["Mon 3"]))
    if data[10] is not None or data[11] is not None or data[12] > 0:
      if data[10] is None:
        issue_message += "/4th monitor needs a make"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Make_4"]))
      if data[11] is None:
        issue_message += "/4th monitor needs a model"
        self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["WS_Mon_Mod_4"]))
      if data[12] <= 0:
        issue_message += "/4th monitor needs a size"
        self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["Mon 4"]))
    if issue_message:
      return issue_message
    return None

  def printer_Issues(self, data):
    # This function looks for printer errors
    # data1 = bc data2 = bf data3 = bd data4 = bg data5 = bh
    # bc = type bf = ip bd = queue name bg = make bh = model
    # data comprised of the following:
    # data[0] = row number
    # data[1] = type
    # data[2] = IP
    # data[3] = Queue
    # data[4] = Make
    # data[5] = Model
    issue_message = ""
    if data[1] is None:
      issue_message += "/Missing printer Type"
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["PRNT_Type"]))
    if data[2] is None:
      issue_message += "/Missing printer IP"
      self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["Network Pntr IP"]))
    if data[3] is None:
      issue_message += "/Missing printer Queue Name"
      self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["PRNT_Queue_Name"]))
    if data[4] is None:
      issue_message += "/Missing printer Make"
      self.list_color_cells.append(self.ws.cell(row=data[0], column=self.column_index_map["PRNT_Make"]))
    if data[5] is None:
      issue_message += "/Missing printer Model"
      self.list_color_cells.append(self.ws.cell(row=data[0],column=self.column_index_map["PRNT_Model"]))

    if issue_message:
      return issue_message
    return None

  def find_both_sequence_and_dupe_erros(self, valid_floor_plans, designators):
    #this function checks for both sequence and dupe errors

    pass

  def type_sequence(self, designator_type_domain):
    #this function checks to make sure the device type matches its designator
    #it also checks to make sure the domain is correct
    # parameters of the instance are row, designator, device type, domain
    issue_message = ""
    if designator_type_domain[1] is not None and designator_type_domain[2] is not None:
      if designator_type_domain[1].lower().startswith("wow"):
        pass
      elif designator_type_domain[1].lower().startswith('l'):
        if designator_type_domain[2].lower() not in "laptop":
          issue_message += "/device type doesn't match"
          self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0],column=self.column_index_map["Flr Pln D"]))
          self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0],column=self.column_index_map["WS_Type"]))
      elif designator_type_domain[1].lower().startswith(
          'w') and designator_type_domain[2].lower() not in "desktop":
        issue_message += "/device type doesn't match"
        self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0],column=self.column_index_map["Flr Pln D"]))
        self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0], column=self.column_index_map["WS_Type"]))
        self.list_color_cells.append(self.ws.cell(row=designator_type_domain[0],column=self.column_index_map["Domain"]))
    return issue_message

#end of issue functions----------------------------------------------------------
#####################################################
######################################################
######################################################
######################################################
#implementation functions------------------------------------------------------------

  def flagging_issues(self):
    #this caller function will loop each row and call issue functions
    issue_messages = []
    for i in range(3, self.ws.max_row + 1):

      #Collect all issue messages for the current row
      row_issues = []

      #Check for floorplan issues
      row_issues.append(self.find_missing_floor_names([
              i,self.ws.cell(row=i,column=self.column_index_map["Flr Pln N"]).value,
              self.ws.cell(row=i,column=self.column_index_map["Flr Pln L"]).value
          ]))

      #Check for designator issues
      row_issues.append(self.designatorIssues([
              i,self.ws.cell(row=i,column=self.column_index_map["Flr Pln D"]).value,
              self.ws.cell(row=i, column=self.column_index_map["Type"]).value
          ]))

      #Check for device type issues
      row_issues.append(self.type_sequence([
              i,
              self.ws.cell(row=i,column=self.column_index_map["Flr Pln D"]).value,
              self.ws.cell(row=i,column=self.column_index_map["WS_Type"]).value,
              self.ws.cell(row=i, column=self.column_index_map["Domain"])
          ]))

      #Check for monitor issues
      if "workstation" in self.ws.cell(
          row=i, column=self.column_index_map["Type"]).value.lower():
        params = [
            i,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Make_1"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Mod_1"]).value,
            self.ws.cell(row=i, column=self.column_index_map["Mon 1"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Make_2"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Mod_2"]).value,
            self.ws.cell(row=i, column=self.column_index_map["Mon 2"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Make_3"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Mod_3"]).value,
            self.ws.cell(row=i, column=self.column_index_map["Mon 3"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Make_4"]).value,
            self.ws.cell(row=i,column=self.column_index_map["WS_Mon_Mod_4"]).value,
            self.ws.cell(row=i, column=self.column_index_map["Mon 4"]).value
        ]
        row_issues.append(self.find_monitor_issues(params))

      #Check for printer issues
      if "printer" in self.ws.cell(
          row=i, column=self.column_index_map["Type"]).value.lower():
        row_issues.append(
            self.printer_Issues([
                i,
                self.ws.cell(row=i,column=self.column_index_map["PRNT_Type"]).value,
                self.ws.cell(row=i,column=self.column_index_map["Network Pntr IP"]).value,
                self.ws.cell(row=i,column=self.column_index_map["PRNT_Queue_Name"]).value,
                self.ws.cell(row=i,column=self.column_index_map["PRNT_Make"]).value,
                self.ws.cell(row=i,column=self.column_index_map["PRNT_Model"]).value
            ]))

      # Check for unique Queues and IPs

      #Collect issue messages for the current row
      issue_messages.append(row_issues)

    #Combine issue messages into one string for each row
    for i in range(3, self.ws.max_row + 1):
      combined_message = " ".join(msg for msg in issue_messages[i - 3] if msg)
      self.ws.cell(row=i, column=self.issueColumn).value += combined_message

    return None

  def count_devices(self, valid_floor_names):
    # Initialize device-type counters
    counts = {"Printers": 0, "Workstations": 0, "Laptops": 0, "WOWs": 0, "Specialty Printers": 0}

    # Manually specify the order of the device subtypes to ensure alphabetical output
    ordered_subtypes = ['Laptops', 'Printers', 'Specialty Printers', 'Workstations', 'WOWs']


    # Process each row to count device types
    for i in range(3, self.ws.max_row + 1):
        device_type = self.ws.cell(row=i, column=self.column_index_map["Type"]).value.lower()
        flr_pln_d = self.ws.cell(row=i, column=self.column_index_map["Flr Pln D"]).value

        # Determine device subtype based on Flr Pln D or other criteria
        # This involves analyzing designator formats to correctly classify the device
        # Add conditional checks based on the specifics of designator formats
        
        # Example condition to differentiate device subtypes
        if flr_pln_d is not None:
          if flr_pln_d.startswith("L"):
              counts["Laptops"] += 1
          # Add more conditions here to handle other subtypes
          if flr_pln_d.startswith("W"):
              counts["Workstations"] += 1
          if flr_pln_d.startswith("P"):
              counts["Printers"] += 1
          if flr_pln_d.startswith("S"):
              counts["Specialty Printers"] += 1
          if flr_pln_d.startswith('WOW', 0, 3):
              counts["WOWs"] += 1

    # Format the output to include subtype name before the count
    counts_output = []
    for subtype in ordered_subtypes:
        if subtype in counts:
            counts_output.append(f"{subtype}: {counts[subtype]}")

    # Correctly format the output to include subtype name before the count, in alphabetical order
    counts_str = '\t'.join(f"{subtype}: {counts[subtype]}" for subtype in ordered_subtypes if subtype in counts)

    # Use the first row because it's never touched by processing
    output_row = 1

    # Define a list to map device types to their designated output columns (1st)
    output_column = 1

    # Write the concatenated counts string to the target cell
    self.ws.cell(row=output_row, column=output_column).value = counts_str

    # Ensure this method is called from an appropriate place within the script, with the correct parameters

    # self.output_counts()

    # Compute count for each floor:
    for category in valid_floor_names:
      for floor_plan_name in valid_floor_names[category]:
          self.test_designator_sequence(floor_plan_name)


    for floor_name in valid_floor_names:
      # Accumulate the counts for each device subtype
      counts = {"Printers": 0, "Workstations": 0, "Laptops": 0, "WOWs": 0, "Specialty Printers": 0}
      # Process each row to count device types
      for i in range(3, self.ws.max_row + 1):
        device_type = self.ws.cell(row=i, column=self.column_index_map["Type"]).value.lower()
        flr_pln_d = self.ws.cell(row=i, column=self.column_index_map["Flr Pln D"]).value

        # Determine device subtype based on Flr Pln D or other criteria
        # This involves analyzing designator formats to correctly classify the device
        # Add conditional checks based on the specifics of designator formats
        # Example condition to differentiate device subtypes
        if flr_pln_d is not None:
          if flr_pln_d.startswith("L"):
              counts["Laptops"] += 1
          # Add more conditions here to handle other subtypes
          if flr_pln_d.startswith("W"):
              counts["Workstations"] += 1
          if flr_pln_d.startswith("P"):
              counts["Printers"] += 1
          if flr_pln_d.startswith("S"):
              counts["Specialty Printers"] += 1
          if flr_pln_d.startswith('wow', 0, 3):
              counts["WOWs"] += 1
      return counts
# -------------------------------------------------------
# -------------------------------------------------------
  def test_designator_sequence(self, floor_plan_name):
    # This functions iterates over the designators of a given floor name
    # It will flag issues for designators, pass the cell to be highlighted,
    # and output the record ID number of the designator, header "ID"
    print("\n\nTesting sequences for floor plan:", floor_plan_name)
    designators = []
    seen_designators = {}  # To track seen designators and their rows
    last_seen = {}  # To track the last seen number for each prefix

    print("\nStarting the for loop to gather unsorted designators")
    
    for i in range(3, self.ws.max_row + 1):
        flr_pln_n = self.ws.cell(row=i, column=self.column_index_map["Flr Pln N"]).value
        record_id = self.ws.cell(row=i, column=self.column_index_map["ID"])
        # Make sure the floor name of the current row being processed
        # store the record ID number of the designator
        # matches the valid floor name passed in
        if flr_pln_n == floor_plan_name:
            designator = self.ws.cell(row=i, column=self.column_index_map["Flr Pln D"]).value
            if designator:
              if flr_pln_n == floor_plan_name:  # Add debugging here
                print(f"\nCollecting designator for floor plan {flr_pln_n}: {designator}")
                designators.append((designator, i))  # Include row number for reference

    print("\nFinished the for loop to gather unsorted designators")
    print(f"List of unsorted designators: {designators}")
    

    print("\nStarting the for loop to sort designators")
    # Sort the designators, adjusting for 'WOW'
    sorted_designators = sorted(designators, key=lambda x: (
        x[0][0:3] if x[0].startswith("WOW") else x[0][0], 
        int(''.join(filter(str.isdigit, x[0][3:])) if x[0].startswith("WOW") else ''.join(filter(str.isdigit, x[0][1:]))) if ''.join(filter(str.isdigit, x[0][3:]) if x[0].startswith("WOW") else ''.join(filter(str.isdigit, x[0][1:]))) else 0
    ))
    print("\nFinished the for loop to sort designators")

    # Print the sorted designators
    print(f"\nSorted designators: {sorted_designators}")
    # Print the type of this variable
    print(f"Type of sorted_designators: {type(sorted_designators)}")

    # Loop for duplicates and sequence errors
    print("\nStarting the for loop to check for duplicates and sequence errors")
    for designator, row in designators:
        prefix = "WOW" if designator.startswith("WOW") else designator[0]
      
      # Use a conditional expression to handle empty strings after filtering for digits
        num_part_str = ''.join(filter(str.isdigit, designator))
        num_part = int(num_part_str) if num_part_str else 0  # Convert to int if not empty, else default to 0
        if prefix in seen_designators:
            if num_part == seen_designators[prefix][0]:
                # Duplicate found
                original_row = seen_designators[prefix][1]
                self.flag_issue(row, f"/is duplicate with {original_row}")
            elif num_part != seen_designators[prefix][0] + 1:
                # Skip detected
                self.flag_issue(row, "/skip in sequence")
        seen_designators[prefix] = (num_part, row)  # Update with current num and row

    print()
    
    # Print the seen_designators dictionary
    print("\nSeen designators dictionary:")
    for des in seen_designators:
        print(f"{des}: {seen_designators[des]}")
    print(seen_designators)  
    
  def flag_issue(self, row, message):
    current_value = self.ws.cell(row=row, column=self.column_index_map["Issues"]).value
    if current_value:
        self.ws.cell(row=row, column=self.column_index_map["Issues"]).value = current_value + message
    else:
        self.ws.cell(row=row, column=self.column_index_map["Issues"]).value = message

  # -------------------------------------------------------
  # -------------------------------------------------------

  
  def almost_test_designator_sequence(self, floor_plan_name):
    print("start of testing sequences")
    # Initialize a dictionary to track designators and their rows
    designators = []

    # Collect designators for the specified floor plan
    for i in range(3, self.ws.max_row + 1):
        flr_pln_n = self.ws.cell(row=i, column=self.column_index_map["Flr Pln N"]).value
        if flr_pln_n == floor_plan_name:
            designator = self.ws.cell(row=i, column=self.column_index_map["Flr Pln D"]).value
            if designator:
                designators.append(designator)

    # Sort the designators
    designators.sort(key=lambda x: (x[0], int(x[1:])) if x[0] in 'LPSW' and x[1:].isdigit() else (x, 0))

    # Remember to modify the rest of the function as needed to work with the sorted designators

  def old_test_designator_sequence(self, floor_plan_names):
    # Initialize a dictionary to track the last seen number for each designator type
    last_seen = {"L": 0, "P": 0, "S": 0, "W": 0, "WOW": 0}
    # Initialize a dictionary to track skips for each designator type
    skips = {"L": [], "P": [], "S": [], "W": [], "WOW": []}

    for i in range(3, self.ws.max_row + 1):
        flr_pln_n = self.ws.cell(row=i, column=self.column_index_map["Flr Pln N"]).value
        if flr_pln_n == floor_plan_names:
            designator = self.ws.cell(row=i, column=self.column_index_map["Flr Pln D"]).value
            if designator:
                # Extract the prefix and numeric part of the designator
                prefix = designator[0]
                if prefix in last_seen:
                    try:
                        number = int(designator[1:])  # Assume numeric part follows the prefix
                        if prefix == 'W' and designator.startswith("WOW"):
                            prefix = "WOW"  # Special case for WOW designators
                            number = int(designator[3:])  # Adjust for the WOW prefix
                        # Check for skips in the sequence
                        if number != last_seen[prefix] + 1:
                            # Record the skipped numbers
                            skips[prefix].extend(range(last_seen[prefix] + 1, number))
                        last_seen[prefix] = number
                    except ValueError:
                        # Handle cases where the designator does not follow the expected format
                        pass

    # Output the skips for review
    for prefix, skip_list in skips.items():
        if skip_list:
            print(f"Skips for {prefix} in {floor_plan_name}: {skip_list}")
        else:
            print(f"No skips for {prefix} in {floor_plan_name}")

    return skips

  def old_output_counts(self):
    #This logic will output devices counts to the cells beneath the data

        total_label = "Total Devices"
        start_row = self.ws.max_row + 5
        start_col = self.column_index_map[total_label]
        self.ws.cell(row=start_row, column=start_col).value = total_label
        column_letters = list('CDEFGHIJ')
        for i, (key, value) in enumerate(counts.items()):
          col = start_col + i + 1
          label = self.ws.cell(row=1, column=col).value
          self.ws.cell(row=start_row, column=col).value = value
          self.ws.cell(row=start_row +1 , column=col).value = f"=SUM({column_letters[i]}3:{column_letters[i]}{start_row})"


  def process_data(self):
    print(f"Beginning of process_data: {self.output_file}")
    self.setup_logging()
    self.load_Data()
    self.validate_floor_plans()
    self.flagging_issues()
    self.highlight_rows()
    self.save_output_file()
    print(f"Ending of process_data: {self.output_file}")
    return None


# -------------------------------------------------
# End of Processing Functions
# -------------------------------------------------

# -------------------------------------------------
# Start Implementation Functions------------------------------------------------------------
# -------------------------------------------------

# Instantiate the DataProcessor class
print("Starting the data the processor")
data_processor = DataProcessor()

# Call the process_data method on the instance
data_processor.process_data()

print("After calling the process_data method on the instance")
# Configure logging to write messages/logs to a file
logging.basicConfig(filename='data_processing.log', level=print)

# -------------------------------------------------
# End of Implementation Functions
# -------------------------------------------------
